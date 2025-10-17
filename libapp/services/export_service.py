"""Service d'exportation des données vers des formats de fichiers standards.

Ce module fournit des fonctions pour exporter des listes de données
(livres, membres, etc.) vers des fichiers CSV ou XLSX avec métadonnées.
"""

from __future__ import annotations

import csv
from collections.abc import Sequence
from datetime import datetime
from pathlib import Path
from typing import Any, Literal

from openpyxl import Workbook
from openpyxl.styles import Font

from ..services.translation_service import translate
from .metrics_service import benchmark

# 📋 Type alias pour clarté
ExportFormat = Literal["csv", "xlsx"]


class ExportMetadata:
    """Métadonnées à inclure dans un export.

    Attributes:
        include_date: Inclure la date d'export.
        include_count: Inclure le nombre d'enregistrements.
        include_custom_message: Inclure un message personnalisé.
        custom_message: Le message personnalisé à inclure.
        app_name: Nom de l'application (si configuré).
        library_name: Nom de la bibliothèque (si configuré).
    """

    def __init__(
        self,
        include_date: bool = True,
        include_count: bool = True,
        include_custom_message: bool = False,
        custom_message: str = "",
        app_name: str = "Biblio",
        library_name: str = "",
    ):
        """Initialise les métadonnées d'export.

        Args:
            include_date: Si True, inclut la date d'export.
            include_count: Si True, inclut le nombre d'enregistrements.
            include_custom_message: Si True, inclut un message personnalisé.
            custom_message: Le message personnalisé (si activé).
            app_name: Nom de l'application.
            library_name: Nom de la bibliothèque (optionnel).
        """
        self.include_date = include_date
        self.include_count = include_count
        self.include_custom_message = include_custom_message
        self.custom_message = custom_message
        self.app_name = app_name
        self.library_name = library_name

    def generate_lines(self, record_count: int) -> list[str]:
        """Génère les lignes de métadonnées à inclure dans l'export.

        Args:
            record_count: Nombre d'enregistrements exportés.

        Returns:
            Liste de strings représentant les métadonnées.
        """
        lines = []

        # Ligne app/library name
        if self.library_name:
            lines.append(f"{self.app_name} - {self.library_name}")
        else:
            lines.append(self.app_name)

        # Date d'export
        if self.include_date:
            export_date = datetime.now().strftime("%d/%m/%Y %H:%M")
            lines.append(f"{translate('export.metadata.date')}: {export_date}")

        # Nombre d'enregistrements
        if self.include_count:
            lines.append(f"{translate('export.metadata.count')}: {record_count}")

        # Message personnalisé
        if self.include_custom_message and self.custom_message:
            lines.append(self.custom_message)

        return lines


@benchmark("export_csv")
def export_data_to_csv(
    filepath: Path,
    headers: list[str],
    data: Sequence[Sequence[Any]],
    metadata: ExportMetadata | None = None,
) -> None:
    """Exporte des données vers un fichier CSV.

    Args:
        filepath: Chemin du fichier de destination.
        headers: Liste des en-têtes de colonnes.
        data: Données à exporter (liste de listes/tuples).
        metadata: Métadonnées optionnelles à inclure en début de fichier.

    Raises:
        IOError: Si l'écriture dans le fichier échoue.
    """
    with filepath.open("w", newline="", encoding="utf-8-sig") as csvfile:
        writer = csv.writer(csvfile, delimiter=";")

        # Écrire les métadonnées si présentes
        if metadata:
            meta_lines = metadata.generate_lines(len(data))
            for line in meta_lines:
                writer.writerow([line])
            # Ligne vide de séparation
            writer.writerow([])

        # Écrire les en-têtes
        writer.writerow(headers)

        # Écrire les données
        for row in data:
            writer.writerow(row)


@benchmark("export_xlsx")
def export_data_to_xlsx(
    filepath: Path,
    headers: list[str],
    data: Sequence[Sequence[Any]],
    metadata: ExportMetadata | None = None,
    sheet_name: str = "Export",
) -> None:
    """Exporte des données vers un fichier Excel XLSX.

    Args:
        filepath: Chemin du fichier de destination.
        headers: Liste des en-têtes de colonnes.
        data: Données à exporter (liste de listes/tuples).
        metadata: Métadonnées optionnelles à inclure en début de fichier.
        sheet_name: Nom de la feuille Excel (par défaut "Export").

    Raises:
        IOError: Si l'écriture dans le fichier échoue.
    """
    workbook = Workbook()
    sheet = workbook.active
    if sheet is None:
        msg = "Unable to create Excel worksheet"
        raise RuntimeError(msg)
    sheet.title = sheet_name

    current_row = 1

    # Écrire les métadonnées si présentes
    if metadata:
        meta_lines = metadata.generate_lines(len(data))
        for line in meta_lines:
            cell = sheet.cell(row=current_row, column=1, value=line)
            # Style en gras et italique pour les métadonnées
            cell.font = Font(bold=True, italic=True)
            current_row += 1
        # Ligne vide de séparation
        current_row += 1

    # Écrire les en-têtes
    for col_idx, header in enumerate(headers, start=1):
        cell = sheet.cell(row=current_row, column=col_idx, value=header)
        # En-têtes en gras
        cell.font = Font(bold=True)

    current_row += 1

    # Écrire les données
    for row_data in data:
        for col_idx, value in enumerate(row_data, start=1):
            sheet.cell(row=current_row, column=col_idx, value=value)
        current_row += 1

    # Sauvegarder le fichier
    workbook.save(filepath)


def export_data(
    filepath: Path,
    headers: list[str],
    data: Sequence[Sequence[Any]],
    file_format: ExportFormat,
    metadata: ExportMetadata | None = None,
    sheet_name: str = "Export",
) -> None:
    """Exporte des données vers CSV ou XLSX selon le format spécifié.

    Fonction générique qui délègue à la fonction d'export appropriée.

    Args:
        filepath: Chemin du fichier de destination.
        headers: Liste des en-têtes de colonnes.
        data: Données à exporter (liste de listes/tuples).
        file_format: Format d'export ("csv" ou "xlsx").
        metadata: Métadonnées optionnelles à inclure.
        sheet_name: Nom de la feuille (pour XLSX uniquement).

    Raises:
        ValueError: Si le format est invalide.
        IOError: Si l'écriture échoue.
    """
    if file_format == "csv":
        export_data_to_csv(filepath, headers, data, metadata)
    elif file_format == "xlsx":
        export_data_to_xlsx(filepath, headers, data, metadata, sheet_name)
    else:
        msg = f"Format non supporté : {file_format}"
        raise ValueError(msg)
