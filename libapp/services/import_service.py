"""
Service de gestion de l'import de données depuis des fichiers Excel (XLSX).

Ce module fournit un pipeline complet pour l'import de livres :
1.  **Parsing** (`parse_xlsx`): Lit un fichier XLSX et le transforme en une
    liste de dictionnaires Python, un par ligne.
2.  **Validation** (`validate_rows`): Vérifie la conformité des données lues
    (champs obligatoires, types, etc.).
3.  **Upsert** (`upsert_rows`): Insère ou met à jour les données dans la base
    de données en appliquant une politique de gestion des doublons
    (skip, merge, replace).

Il est conçu pour être utilisé par l'interface utilisateur d'import, en lui
fournissant des retours sur l'avancement et les erreurs.
"""

from __future__ import annotations

import csv
import logging
from collections.abc import Iterator
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Literal, Protocol

import openpyxl
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from ..persistence.database import get_session  # aligne avec projet courant
from ..persistence.models_sa import Author, Book  # aligne avec projet courant
from .types import ImportBatch, ImportErrorItem, ImportResult

try:
    # Nom actuel
    from .utils import clean_author as _clean_author
except Exception:
    try:
        # Nom legacy
        from .utils import cleanauthor as _clean_author  # type: ignore[attr-defined]
    except Exception:

        def _clean_author(s: str | None) -> str:
            """Fallback minimal si util non dispo."""
            return " ".join((s or "").split()).strip().title()


try:
    # Nom actuel
    from .utils import normalize_isbn as _normalize_isbn
except Exception:
    try:
        # Nom legacy
        from .utils import normalizeisbn as _normalize_isbn  # type: ignore[attr-defined]
    except Exception:

        def _normalize_isbn(val: str | None) -> str | None:
            """Fallback: garde chiffres (+ X), retourne None si invalide."""
            if not val:
                return None
            raw = str(val)
            digits = "".join(ch for ch in raw if ch.isdigit() or ch.upper() == "X")
            if len(digits) in (10, 13):
                return digits
            return digits or None


# --- Conversion des champs internes (UI EN -> parseur FR) ---
_FR_REQUIRED = {"titre"}
_FR_OPTIONAL = {
    "isbn",
    "sous_titre",
    "auteurs",
    "editeur",
    "date_publication",
    "collection",
    "tome",
    "code_interne",
    "mots_cles",
}

_EN_TO_FR = {
    "title": "titre",
    "subtitle": "sous_titre",
    "author": "auteurs",
    "authors": "auteurs",
    "isbn": "isbn",
    "publisher": "editeur",
    "year": "date_publication",
    "collection": "collection",
    "fund": "collection",
    "volume": "tome",
    "tome": "tome",
    "code": "code_interne",
    "tags": "mots_cles",
}


class MappingError(Exception):
    """Erreur levée lorsque le mappage des colonnes échoue."""

    pass


class BaseImporter(Protocol):
    """Protocole définissant l'interface pour tous les importeurs."""

    def extract_rows(self) -> Iterator[dict[str, Any]]: ...

    def map_rows(self, raw_rows: list[dict]) -> tuple[list[dict], dict]: ...


class CsvImporter:
    """Importeur pour les fichiers CSV."""

    def extract_rows(self, file_path: Path) -> Iterator[dict[str, Any]]:
        with file_path.open(mode="r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            yield from reader

    def map_rows(self, raw_rows: list[dict]) -> tuple[list[dict], dict]:
        # Logique de mappage simplifiée pour le moment
        return raw_rows, {}


class ExcelImporter:
    """Importeur pour les fichiers Excel."""

    def extract_rows(self, file_path: Path) -> Iterator[dict[str, Any]]:
        workbook = openpyxl.load_workbook(file_path, read_only=True)
        sheet = workbook.active
        headers = [cell.value for cell in sheet[1]]
        for row in sheet.iter_rows(min_row=2):
            yield {headers[i]: cell.value for i, cell in enumerate(row)}

    def map_rows(self, raw_rows: list[dict]) -> tuple[list[dict], dict]:
        # Logique de mappage simplifiée
        return raw_rows, {}


logger = logging.getLogger(__name__)

# ImportPolicy sera défini dans types.py


def get_importer_for_file(file_path: Path) -> BaseImporter:
    """Retourne l'instance d'importer appropriée pour le type de fichier."""
    extension = file_path.suffix.lower()
    if extension == ".csv":
        return CsvImporter()
    elif extension in [".xlsx", ".xls"]:
        return ExcelImporter()
    else:
        raise ValueError(f"Type de fichier non supporté: {extension}")


def _normalize_user_mapping(user_mapping: dict[str, str]) -> dict[str, str]:
    """
    Convertit le mapping UI {champ_interne_EN: header_xlsx}
    en mapping parseur {header_xlsx: champ_interne_FR}.
    """
    normalized: dict[str, str] = {}
    for internal_en, colname in user_mapping.items():
        field_fr = _EN_TO_FR.get(internal_en, internal_en)
        if field_fr in (_FR_REQUIRED | _FR_OPTIONAL) and colname:
            normalized[colname] = field_fr
    return normalized


# --- Parsing XLSX -> ImportBatch ---
def parse_xlsx(file_path: str | Path, mapping_ui: dict[str, str]) -> ImportBatch:
    """
    Lit un XLSX et retourne (rows normalisées, erreurs).
    mapping_ui: {champ_interne_EN: 'Header XLSX'}.
    """
    mapping = _normalize_user_mapping(mapping_ui)
    wb = load_workbook(filename=str(file_path), read_only=True, data_only=True)
    ws = wb.active

    header_row = next(ws.iter_rows(min_row=1, max_row=1))
    headers = [
        (
            c.value.strip()
            if isinstance(c.value, str)
            else (str(c.value) if c.value is not None else "")
        )
        for c in header_row
    ]

    errors: list[ImportErrorItem] = []
    col_idx_for_field: dict[str, int] = {}
    for header_name, field in mapping.items():
        try:
            idx = headers.index(header_name)
            col_idx_for_field[field] = idx
        except ValueError:
            errors.append(ImportErrorItem(0, field, f"Colonne introuvable: {header_name}", "error"))

    @dataclass(slots=True)
    class BookRow:
        titre: str
        sous_titre: str | None = None
        auteurs: list[str] | None = None
        isbn: str | None = None
        editeur: str | None = None
        date_publication: int | None = None
        collection: str | None = None
        tome: int | None = None
        code_interne: str | None = None
        mots_cles: list[str] | None = None

    rows: list[BookRow] = []
    for i, r in enumerate(ws.iter_rows(min_row=2), start=2):
        data: dict[str, object] = {}
        for field, idx in col_idx_for_field.items():
            v = r[idx].value
            data[field] = v.strip() if isinstance(v, str) else v

        titre = (str(data.get("titre") or "")).strip()
        if not titre:
            errors.append(ImportErrorItem(i, "titre", "Titre manquant", "error"))
            continue

        # auteurs
        auteurs_list: list[str] | None = None
        if data.get("auteurs") not in (None, ""):
            raw = str(data.get("auteurs"))
            parts = [p.strip() for p in raw.replace(";", ",").split(",")]
            cleaned = [_clean_author(p) for p in parts if p.strip()]
            seen: set[str] = set()
            auteurs_list = []
            for n in cleaned:
                if n and n not in seen:
                    seen.add(n)
                    auteurs_list.append(n)

        # isbn
        isbn = _normalize_isbn(str(data.get("isbn")) if data.get("isbn") else None)

        # tome
        tome = None
        if data.get("tome") not in (None, ""):
            try:
                tome = int(data["tome"])
            except Exception:
                errors.append(
                    ImportErrorItem(i, "tome", f"Tome non entier: {data.get('tome')}", "warning")
                )
                tome = None

        # mots_cles
        mots_cles = None
        if data.get("mots_cles") not in (None, ""):
            mots_cles = [s.strip() for s in str(data.get("mots_cles")).split(";") if s.strip()]

        # année
        year = None
        if data.get("date_publication") not in (None, ""):
            try:
                year = int(str(data["date_publication"]))
            except Exception:
                year = None

        rows.append(
            BookRow(
                titre=titre,
                sous_titre=(
                    str(data.get("sous_titre")).strip() if data.get("sous_titre") else None
                ),
                auteurs=auteurs_list,
                isbn=isbn,
                editeur=(str(data.get("editeur")).strip() if data.get("editeur") else None),
                date_publication=year,
                collection=(
                    str(data.get("collection")).strip() if data.get("collection") else None
                ),
                tome=tome,
                code_interne=(
                    str(data.get("code_interne")).strip() if data.get("code_interne") else None
                ),
                mots_cles=mots_cles,
            )
        )

    return ImportBatch(rows=rows, errors=errors)


# --- Validation douce ---
def validate_rows(rows: list, _progress=None) -> tuple[list, list[ImportErrorItem]]:
    warnings: list[ImportErrorItem] = []
    out: list = []
    for idx, r in enumerate(rows, start=2):
        # isbn normalisé + longueur prudente
        if getattr(r, "isbn", None):
            r.isbn = _normalize_isbn(r.isbn)
            if r.isbn and len(r.isbn) not in (10, 13):
                warnings.append(
                    ImportErrorItem(idx, "isbn", f"Longueur ISBN suspecte: {r.isbn}", "warning")
                )
        out.append(r)
        if _progress:
            _progress(idx, len(rows), "Validation…")
    return out, warnings


# --- Upsert avec politiques ---
def _get_author_objs(session: Session, names: list[str]) -> list[Author]:
    objs: list[Author] = []
    for name in names:
        a = session.query(Author).filter_by(name=name).one_or_none()
        if not a:
            a = Author(name=name)
            session.add(a)
            session.flush()
        objs.append(a)
    return objs


def upsert_rows(
    rows: list, on_isbn_conflict: Literal["skip", "merge", "replace"] = "merge", _progress=None
) -> ImportResult:
    inserted = updated = skipped = 0
    errors: list[ImportErrorItem] = []
    with get_session() as session:
        for i, r in enumerate(rows, start=2):
            try:
                existing = None
                if getattr(r, "isbn", None):
                    existing = session.execute(
                        select(Book).where(Book.isbn == r.isbn)
                    ).scalar_one_or_none()
                if existing is None:
                    # insert
                    authors_list = [a for a in (r.auteurs or []) if a]
                    b = Book(
                        title=r.titre,
                        subtitle=getattr(r, "sous_titre", None),
                        isbn=r.isbn,
                        publisher=getattr(r, "editeur", None),
                        year=getattr(r, "date_publication", None),
                        collection=getattr(r, "collection", None),
                        volume=getattr(r, "tome", None),
                        code_interne=getattr(r, "code_interne", None),
                        mots_cles=(
                            ";".join(r.mots_cles) if getattr(r, "mots_cles", None) else None
                        ),
                        authors_text=", ".join(authors_list) or None,
                    )
                    session.add(b)
                    session.flush()
                    if authors_list:
                        b.authors = _get_author_objs(session, authors_list)
                    inserted += 1
                else:
                    if on_isbn_conflict == "skip":
                        skipped += 1
                        continue
                    elif on_isbn_conflict == "replace":
                        existing.title = r.titre
                        existing.subtitle = getattr(r, "sous_titre", None)
                        existing.isbn = r.isbn
                        existing.publisher = getattr(r, "editeur", None)
                        existing.year = getattr(r, "date_publication", None)
                        existing.collection = getattr(r, "collection", None)
                        existing.volume = getattr(r, "tome", None)
                        existing.code_interne = getattr(r, "code_interne", None)
                        existing.mots_cles = (
                            ";".join(r.mots_cles) if getattr(r, "mots_cles", None) else None
                        )
                        authors_list = [a for a in (r.auteurs or []) if a]
                        existing.authors_text = ", ".join(authors_list) or None
                        existing.authors = _get_author_objs(session, authors_list)
                        updated += 1
                    else:
                        # merge doux
                        def choose(cur, new, join=False):
                            if cur in (None, "") and new not in (None, "", []):
                                return ";".join(new) if join else new
                            return cur

                        existing.subtitle = choose(
                            existing.subtitle, getattr(r, "sous_titre", None)
                        )
                        existing.isbn = choose(existing.isbn, r.isbn)
                        existing.publisher = choose(existing.publisher, getattr(r, "editeur", None))
                        existing.year = choose(existing.year, getattr(r, "date_publication", None))
                        existing.collection = choose(
                            existing.collection, getattr(r, "collection", None)
                        )
                        existing.volume = choose(existing.volume, getattr(r, "tome", None))
                        existing.code_interne = choose(
                            existing.code_interne, getattr(r, "code_interne", None)
                        )
                        existing.mots_cles = choose(
                            existing.mots_cles, getattr(r, "mots_cles", None), join=True
                        )
                        authors_list = [a for a in (r.auteurs or []) if a]
                        if not existing.authors_text and authors_list:
                            existing.authors_text = ", ".join(authors_list) or None
                            existing.authors = _get_author_objs(session, authors_list)
                        updated += 1
            except Exception as e:
                errors.append(ImportErrorItem(i, "row", f"{type(e).__name__}: {e}", "error"))
            if _progress:
                _progress(i, len(rows), "Enregistrement…")
        try:
            session.commit()
        except Exception:
            session.rollback()
            raise
    return ImportResult(inserted=inserted, updated=updated, skipped=skipped, errors=errors)


# ============================================================================
# 🔥 IMPORT MEMBRES (version simplifiée)
# ============================================================================


@dataclass(slots=True)
class MemberRow:
    """Structure temporaire pour un membre importé."""

    member_no: str
    first_name: str
    last_name: str
    email: str | None = None
    phone: str | None = None
    status: str = "apprenti"
    is_active: bool = True


# Mapping des champs membres (EN → FR interne)
_MEMBER_EN_TO_FR = {
    "member_no": "numero_membre",
    "first_name": "prenom",
    "last_name": "nom",
    "email": "email",
    "phone": "telephone",
    "status": "statut",
    "is_active": "actif",
}

_FR_MEMBER_REQUIRED = {"numero_membre", "prenom", "nom"}
_FR_MEMBER_OPTIONAL = {"email", "telephone", "statut", "actif"}


def parse_members_xlsx(file_path: str | Path, mapping_ui: dict[str, str]) -> ImportBatch:
    """
    Lit un fichier XLSX/CSV de membres et retourne les données normalisées.

    Args:
        file_path: Chemin du fichier XLSX ou CSV
        mapping_ui: Mapping {champ_interne_EN: 'Header fichier'}

    Returns:
        ImportBatch avec rows (MemberRow) et erreurs
    """
    # Normaliser le mapping
    mapping = {}
    for internal_en, colname in mapping_ui.items():
        field_fr = _MEMBER_EN_TO_FR.get(internal_en, internal_en)
        if field_fr in (_FR_MEMBER_REQUIRED | _FR_MEMBER_OPTIONAL) and colname:
            mapping[colname] = field_fr

    # Lire le fichier
    wb = load_workbook(filename=str(file_path), read_only=True, data_only=True)
    ws = wb.active

    # Extraire les headers
    header_row = next(ws.iter_rows(min_row=1, max_row=1))
    headers = [
        c.value.strip() if isinstance(c.value, str) else (str(c.value) if c.value else "")
        for c in header_row
    ]

    errors: list[ImportErrorItem] = []
    col_idx_for_field: dict[str, int] = {}

    # Mapper les colonnes
    for header_name, field in mapping.items():
        try:
            idx = headers.index(header_name)
            col_idx_for_field[field] = idx
        except ValueError:
            errors.append(ImportErrorItem(0, field, f"Colonne introuvable: {header_name}", "error"))

    rows: list[MemberRow] = []

    # Parser les lignes
    for i, r in enumerate(ws.iter_rows(min_row=2), start=2):
        data: dict[str, any] = {}
        for field, idx in col_idx_for_field.items():
            v = r[idx].value
            data[field] = v.strip() if isinstance(v, str) else v

        # Champs requis
        numero = (str(data.get("numero_membre") or "")).strip()
        prenom = (str(data.get("prenom") or "")).strip()
        nom = (str(data.get("nom") or "")).strip()

        if not numero:
            errors.append(ImportErrorItem(i, "numero_membre", "Numéro membre manquant", "error"))
            continue
        if not prenom:
            errors.append(ImportErrorItem(i, "prenom", "Prénom manquant", "error"))
            continue
        if not nom:
            errors.append(ImportErrorItem(i, "nom", "Nom manquant", "error"))
            continue

        # Champs optionnels
        email = (str(data.get("email") or "")).strip() or None
        phone = (str(data.get("telephone") or "")).strip() or None
        status = (str(data.get("statut") or "apprenti")).strip().lower()

        # Valider le statut
        if status not in ("apprenti", "compagnon", "maitre"):
            errors.append(ImportErrorItem(i, "statut", f"Statut invalide: {status}", "warning"))
            status = "apprenti"

        # Actif (boolean)
        is_active = True
        actif_val = data.get("actif")
        if actif_val is not None:
            if isinstance(actif_val, bool):
                is_active = actif_val
            elif isinstance(actif_val, str):
                is_active = actif_val.strip().lower() in ("oui", "yes", "true", "1")
            else:
                is_active = bool(actif_val)

        rows.append(
            MemberRow(
                member_no=numero,
                first_name=prenom,
                last_name=nom,
                email=email,
                phone=phone,
                status=status,
                is_active=is_active,
            )
        )

    return ImportBatch(rows=rows, errors=errors)


def upsert_members(
    rows: list[MemberRow], on_conflict: Literal["skip", "update"] = "skip", _progress=None
) -> ImportResult:
    """
    Insère ou met à jour les membres dans la base.

    Args:
        rows: Liste de MemberRow à importer
        on_conflict: Politique si member_no existe déjà ("skip" ou "update")
        _progress: Callback optionnel pour la progression

    Returns:
        ImportResult avec statistiques (inserted, updated, skipped, errors)
    """
    from ..persistence.models_sa import Member, MemberStatus

    inserted = updated = skipped = 0
    errors: list[ImportErrorItem] = []

    with get_session() as session:
        for i, r in enumerate(rows, start=2):
            try:
                # Chercher si le membre existe déjà
                existing = session.execute(
                    select(Member).where(Member.member_no == r.member_no)
                ).scalar_one_or_none()

                if existing is None:
                    # Insérer nouveau membre
                    m = Member(
                        member_no=r.member_no,
                        first_name=r.first_name,
                        last_name=r.last_name,
                        email=r.email or "",
                        phone=r.phone,
                        status=MemberStatus(r.status),
                        is_active=r.is_active,
                    )
                    session.add(m)
                    inserted += 1
                else:
                    # Membre existe déjà
                    if on_conflict == "skip":
                        skipped += 1
                        continue
                    elif on_conflict == "update":
                        # Mettre à jour
                        existing.first_name = r.first_name
                        existing.last_name = r.last_name
                        if r.email:
                            existing.email = r.email
                        if r.phone:
                            existing.phone = r.phone
                        existing.status = MemberStatus(r.status)
                        existing.is_active = r.is_active
                        updated += 1

                if _progress:
                    _progress(i, len(rows), "Enregistrement...")

            except Exception as e:
                errors.append(ImportErrorItem(i, "row", f"{type(e).__name__}: {e}", "error"))

        try:
            session.commit()
        except Exception:
            session.rollback()
            raise

    return ImportResult(inserted=inserted, updated=updated, skipped=skipped, errors=errors)
